import os

#TODO work out spreadsheet incorporation
def main():

    ss = load_spreadsheet(file_name="../assets/spreadsheets/sheet1.xlsx", sheet_name='Sheet1')

    ss_keys = list(ss[0].keys())

    flag_keys = [k for k in ss_keys if k[0:4]=='flag']

    for row in ss:
        all_flags_true = True
        for f in flag_keys:
            if row[f] != 'x':
                all_flags_true = False

        if all_flags_true:
            print(row['dialogue'])







def load_spreadsheet(file_name, sheet_name):
    from openpyxl import Workbook

    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_name)

    sheet = workbook[sheet_name]
    # for sheet in workbook.worksheets:

    row_count = sheet.max_row
    column_count = sheet.max_column

    header = []
    for col in range(1,column_count+1):
        header.append(sheet.cell(1, col).value)

    row_values = []

    for row in range(2,row_count+1):
        this_row = {}
        for ind,header_name in enumerate(header):
            this_row[header_name] = sheet.cell(row,ind+1).value
        row_values.append(this_row)


    for row in row_values:
        print(row)
    return row_values


    # for row in sheet.iter_rows():
    #     for cell in row:
    #         print(cell.value)








    # sheet = workbook.active

    # sheet["A1"] = "hello"
    # sheet["B1"] = "world!"
    #
    #
    #
    # workbook.save(filename="hello_world.xlsx")

if __name__ == "__main__":
    main()

